"""SQL data source → materialized CSV dataset.

Real end-to-end: the SQL source points at the SAME Postgres the tests run on,
so create/test-connection/execute/materialize all exercise real drivers and a
real query — no mocked transport anywhere.
"""

from pathlib import Path
from urllib.parse import urlparse
from uuid import uuid4

import pytest

from backend.config import settings
from backend.db.connection import execute, query, query_one


def _own_db_config():
    """Connection parts of the test database itself, as a SQL-source payload."""
    u = urlparse(settings.database_url)
    return {
        "host": u.hostname or "127.0.0.1",
        "port": u.port or 5432,
        "database": (u.path or "/postgres").lstrip("/"),
        "username": u.username or "postgres",
        "password": u.password or "",
        "engine": "postgresql",
    }


@pytest.fixture
def sales_table():
    """Throwaway table in the test DB acting as the customer's ERP table."""
    table = f"pytest_erp_sales_{uuid4().hex[:8]}"
    execute(f"""CREATE TABLE {table} (
        fecha DATE NOT NULL,
        sku   TEXT NOT NULL,
        ventas NUMERIC NOT NULL
    )""")
    execute(
        f"INSERT INTO {table} (fecha, sku, ventas) VALUES "
        "('2026-01-01', 'SKU-A', 10), "
        "('2026-01-02', 'SKU-A', 12), "
        "('2026-01-01', '=SUM(A1:A9)', 5)"  # formula-injection attempt
    )
    yield table
    execute(f"DROP TABLE IF EXISTS {table}")


@pytest.fixture
def sql_source(client, auth_headers, sales_table):
    """Connected SQL source pointing at the test database."""
    resp = client.post(
        "/api/v1/data-sources/sql",
        json={"name": f"erp-{uuid4().hex[:6]}", **_own_db_config()},
        headers=auth_headers,
    )
    assert resp.status_code == 200, resp.text
    src = resp.json()["data"]

    resp = client.post(
        f"/api/v1/data-sources/{src['id']}/test-connection", headers=auth_headers,
    )
    assert resp.json()["data"]["ok"] is True, resp.text
    return src


def _materialized_children(tenant_id, source_id):
    return query(
        "SELECT * FROM datasets WHERE tenant_id = %s AND parent_id = %s",
        (tenant_id, source_id),
    )


class TestMaterialize:
    def test_query_result_becomes_csv_dataset(
        self, client, auth_headers, test_tenant, sql_source, sales_table,
    ):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT fecha, sku, ventas FROM {sales_table} ORDER BY fecha"},
            headers=auth_headers,
        )
        assert resp.status_code == 200, resp.text
        ds = resp.json()["data"]

        row = query_one("SELECT * FROM datasets WHERE id = %s", (ds["id"],))
        assert row["tenant_id"] == test_tenant["id"]
        assert row["source_type"] == "file"
        assert row["file_type"] == "csv"
        assert row["parent_id"] == sql_source["id"]
        assert row["row_count"] == 3
        assert row["column_count"] == 3

        csv_path = Path(row["file_path"])
        assert csv_path.exists()
        content = csv_path.read_text(encoding="utf-8")
        assert content.splitlines()[0] == "fecha,sku,ventas"
        assert "SKU-A" in content
        # Formula-injection guard: the malicious cell must be quote-prefixed.
        assert "'=SUM(A1:A9)" in content

        # The query that produced the snapshot is remembered on the source.
        src_row = query_one("SELECT saved_query FROM datasets WHERE id = %s", (sql_source["id"],))
        assert sales_table in src_row["saved_query"]

    def test_materialized_dataset_previews_like_a_file(
        self, client, auth_headers, sql_source, sales_table,
    ):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT fecha, sku, ventas FROM {sales_table}"},
            headers=auth_headers,
        )
        ds_id = resp.json()["data"]["id"]

        resp = client.get(f"/api/v1/data-sources/{ds_id}/preview", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        preview = resp.json()["data"]
        assert preview["columns"] == ["fecha", "sku", "ventas"]
        assert preview["row_count"] == 3

    def test_viewer_denied_and_no_dataset_created(
        self, client, viewer_headers, test_tenant, sql_source, sales_table,
    ):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT * FROM {sales_table}"},
            headers=viewer_headers,
        )
        assert resp.status_code == 403
        assert _materialized_children(test_tenant["id"], sql_source["id"]) == []

    def test_without_query_rejected(self, client, auth_headers, test_tenant, sql_source):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "sql_source_no_saved_query"
        assert _materialized_children(test_tenant["id"], sql_source["id"]) == []

    def test_empty_result_rejected(self, client, auth_headers, test_tenant, sql_source, sales_table):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT * FROM {sales_table} WHERE 1 = 0"},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "sql_result_empty"
        assert _materialized_children(test_tenant["id"], sql_source["id"]) == []

    def test_over_row_cap_refused_not_truncated(
        self, client, auth_headers, test_tenant, sql_source, sales_table, monkeypatch,
    ):
        monkeypatch.setattr(settings, "sql_materialize_max_rows", 2)
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT * FROM {sales_table}"},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "sql_result_too_large"
        assert _materialized_children(test_tenant["id"], sql_source["id"]) == []

    def test_broken_sql_reports_reason(self, client, auth_headers, test_tenant, sql_source):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": "SELECT * FROM table_that_does_not_exist_xyz"},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "sql_query_failed"
        assert _materialized_children(test_tenant["id"], sql_source["id"]) == []

    def test_file_source_cannot_be_materialized(self, client, auth_headers, csv_bytes):
        resp = client.post(
            "/api/v1/data-sources/file",
            files={"file": ("sales.csv", csv_bytes, "text/csv")},
            headers=auth_headers,
        )
        file_src = resp.json()["data"]
        resp = client.post(
            f"/api/v1/data-sources/{file_src['id']}/materialize",
            json={"sql": "SELECT 1"},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "data_source_not_sql"


class TestExportXlsx:
    def test_full_result_downloads_as_workbook(self, client, auth_headers, sql_source, sales_table):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/export-query",
            json={"sql": f"SELECT fecha, sku, ventas FROM {sales_table} ORDER BY fecha"},
            headers=auth_headers,
        )
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "query-result.xlsx" in resp.headers["content-disposition"]

        import io

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        ws = wb["data"]
        rows = [[c.value for c in r] for r in ws.iter_rows()]
        assert rows[0] == ["fecha", "sku", "ventas"]
        assert len(rows) == 4  # header + 3 data rows
        # Excel executes a leading '=' — the malicious cell must be guarded.
        cells = {v for row in rows for v in row}
        assert "'=SUM(A1:A9)" in cells
        assert "=SUM(A1:A9)" not in cells

    def test_export_without_query_rejected(self, client, auth_headers, sql_source):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/export-query",
            json={},
            headers=auth_headers,
        )
        assert resp.status_code == 400
        assert resp.json()["error_code"] == "sql_source_no_saved_query"


def _xlsx_bytes() -> bytes:
    """Small in-memory sales workbook, the shape a real user uploads."""
    import io

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["fecha", "sku", "ventas"])
    for day in range(1, 11):
        ws.append((f"2026-01-{day:02d}", "SKU-X", 5 + day))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExcelUploadWorksLikeCsv:
    """The wizard path and the data-source path must both accept .xlsx."""

    def test_wizard_upload_accepts_xlsx(self, client, auth_headers, test_tenant):
        resp = client.post(
            "/api/v1/datasets",
            files={"file": (
                "ventas.xlsx", _xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        ds = resp.json()["data"]
        row = query_one("SELECT * FROM datasets WHERE id = %s", (ds["id"],))
        assert row["tenant_id"] == test_tenant["id"]
        assert row["file_type"] == "xlsx"
        assert Path(row["file_path"]).exists()

    def test_data_source_upload_accepts_xlsx_and_previews(self, client, auth_headers):
        resp = client.post(
            "/api/v1/data-sources/file",
            files={"file": (
                "ventas.xlsx", _xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=auth_headers,
        )
        assert resp.status_code == 200, resp.text
        src = resp.json()["data"]
        assert src["row_count"] == 10

        resp = client.get(f"/api/v1/data-sources/{src['id']}/preview", headers=auth_headers)
        assert resp.status_code == 200, resp.text
        preview = resp.json()["data"]
        assert preview["columns"] == ["fecha", "sku", "ventas"]
        assert preview["row_count"] == 10


class TestDatasetListExcludesSqlSources:
    def test_wizard_dataset_list_hides_sql_sources_but_shows_materialized(
        self, client, auth_headers, sql_source, sales_table,
    ):
        resp = client.post(
            f"/api/v1/data-sources/{sql_source['id']}/materialize",
            json={"sql": f"SELECT * FROM {sales_table}", "name": "erp snapshot"},
            headers=auth_headers,
        )
        ds_id = resp.json()["data"]["id"]

        resp = client.get("/api/v1/datasets", headers=auth_headers)
        items = resp.json()["data"]["items"]
        ids = {i["id"] for i in items}
        assert sql_source["id"] not in ids
        assert ds_id in ids
        assert all("password_enc" not in str(i.get("sql_config") or "") for i in items)

    def test_get_dataset_never_returns_sql_config(self, client, auth_headers, sql_source):
        resp = client.get(f"/api/v1/datasets/{sql_source['id']}", headers=auth_headers)
        assert resp.status_code == 200
        assert "sql_config" not in resp.json()["data"]

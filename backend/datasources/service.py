"""
Data Source service — file uploads + SQL connections, both stored in `datasets` table.
SQL passwords are encrypted with Fernet (AES-128-CBC + HMAC) keyed from settings.secret_key.
"""

import base64
import hashlib
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import UploadFile

from backend.config import settings
from backend.db.connection import execute, query, query_one, _json
from backend.errors import AppError
from backend.utils.ids import generate_id

log = logging.getLogger(__name__)

ALLOWED_FILE_EXTENSIONS = {".csv", ".xlsx", ".xls", ".parquet", ".json"}
SQL_ENGINES = {"postgresql", "mysql", "mssql", "oracle"}
MAX_FILE_MB = 50

# Every rejection a user can trip in this module answers with 400, the status
# the API's `_service_error` wrapper used before these became AppErrors — the
# wire contract is unchanged, only `error_code` + `error_params` are added.
_REJECTED = 400


def _mb(size_bytes: float) -> float:
    """Bytes as megabytes, rounded for display. A param, never baked into prose."""
    return round(size_bytes / 1024 / 1024, 1)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _check_file_size(file_path: str) -> None:
    size = Path(file_path).stat().st_size
    if size > MAX_FILE_MB * 1024 * 1024:
        raise AppError(
            "data_source_file_too_large_to_read",
            f"File too large ({_mb(size)} MB). Maximum allowed is {MAX_FILE_MB} MB.",
            status_code=_REJECTED,
            params={"size_mb": _mb(size), "max_mb": MAX_FILE_MB},
        )


def _reject_unsupported_extension(suffix: str) -> None:
    """Upload/replace guard. The extension and the allowed set travel as params
    so the frontend can build the sentence in the user's language."""
    if suffix not in ALLOWED_FILE_EXTENSIONS:
        allowed = sorted(ALLOWED_FILE_EXTENSIONS)
        raise AppError(
            "data_source_file_type_unsupported",
            f"File type '{suffix}' is not supported. Allowed: {', '.join(allowed)}",
            status_code=_REJECTED,
            params={"extension": suffix, "allowed": ", ".join(allowed)},
        )


def _reject_oversized_upload(size_bytes: int) -> None:
    """Upload/replace size guard, against the configured upload ceiling."""
    max_bytes = settings.max_upload_size_mb * 1024 * 1024
    if not settings.testing_mode and size_bytes > max_bytes:
        raise AppError(
            "data_source_file_too_large",
            f"File is {_mb(size_bytes)} MB, over the {settings.max_upload_size_mb} MB limit.",
            status_code=_REJECTED,
            params={"size_mb": _mb(size_bytes), "max_mb": settings.max_upload_size_mb},
        )


def _make_sql_engine(cfg: dict, statement_timeout_ms: int = 30_000):
    import sqlalchemy
    conn_str = _build_conn_str(cfg)
    engine_type = cfg.get("engine", "postgresql")
    # Timeout keywords are driver-specific: psycopg2/pymysql take
    # connect_timeout, but pyodbc rejects it (its login timeout is `timeout`).
    if engine_type == "postgresql":
        connect_args: dict = {
            "connect_timeout": 15,
            "options": f"-c statement_timeout={statement_timeout_ms}",
        }
    elif engine_type == "mysql":
        timeout_s = max(1, statement_timeout_ms // 1000)
        connect_args = {
            "connect_timeout": 15,
            "read_timeout": timeout_s,
            "write_timeout": timeout_s,
        }
    elif engine_type == "mssql":
        connect_args = {"timeout": 15}
    else:
        connect_args = {}
    return sqlalchemy.create_engine(conn_str, connect_args=connect_args, pool_pre_ping=True)


def _fernet():
    from cryptography.fernet import Fernet
    key = base64.urlsafe_b64encode(hashlib.sha256(settings.secret_key.encode()).digest())
    return Fernet(key)


def _enc(plain: str) -> str:
    return _fernet().encrypt(plain.encode()).decode()


def _dec(enc: str) -> str:
    try:
        return _fernet().decrypt(enc.encode()).decode()
    except Exception:
        # Fallback for legacy base64-encoded passwords already in the database.
        try:
            return base64.b64decode(enc.encode()).decode()
        except Exception:
            return enc


def _public(row: dict) -> dict:
    """Strip sensitive fields before returning to client."""
    if not row:
        return row
    out = dict(row)
    if out.get("sql_config"):
        cfg = dict(out["sql_config"])
        cfg.pop("password_enc", None)
        out["sql_config"] = cfg
    return out


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ── File sources ───────────────────────────────────────────────────────────────

async def create_file_source(
    tenant_id: str,
    user_id: str,
    file: UploadFile,
    name: Optional[str] = None,
    description: Optional[str] = None,
) -> dict:
    suffix = Path(file.filename or "").suffix.lower()
    _reject_unsupported_extension(suffix)

    content = await file.read()
    size_bytes = len(content)
    _reject_oversized_upload(size_bytes)

    source_id = generate_id("ds")
    from backend.storage import paths
    dst_dir = paths.dataset_dir(tenant_id, source_id)
    dst_dir.mkdir(parents=True, exist_ok=True)
    file_path = dst_dir / f"data{suffix}"
    with open(file_path, "wb") as f:
        f.write(content)

    display_name = name or Path(file.filename or "upload").stem

    # Eagerly count rows so the UI shows a non-zero value immediately
    row_count = None
    col_count = None
    try:
        import io as _io
        if suffix == ".csv":
            row_count = max(0, content.count(b"\n") - 1)
            import csv as _csv
            header = content.split(b"\n", 1)[0].decode("utf-8", "replace")
            if header.strip():
                col_count = len(next(_csv.reader([header])))
        elif suffix in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
            row_count = sum((ws.max_row or 1) - 1 for ws in wb.worksheets)
            col_count  = max((ws.max_column or 0) for ws in wb.worksheets) or None
            wb.close()
        elif suffix == ".parquet":
            import pyarrow.parquet as pq
            meta = pq.read_metadata(_io.BytesIO(content))
            row_count = meta.num_rows
            col_count = meta.num_columns
        elif suffix == ".json":
            import json as _json
            data = _json.loads(content)
            if isinstance(data, list):
                row_count = len(data)
                if data and isinstance(data[0], dict):
                    col_count = len(data[0])
    except Exception:
        pass

    execute(
        """INSERT INTO datasets
           (id, tenant_id, name, description, original_filename, file_type, file_path,
            size_bytes, row_count, column_count, source_type, connection_status, uploaded_by, uploaded_at, updated_at)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'file','connected',%s,NOW(),NOW())""",
        (
            source_id, tenant_id, display_name, description,
            file.filename, suffix.lstrip("."), str(file_path),
            size_bytes, row_count, col_count, user_id,
        ),
    )
    return _public(get_source(tenant_id, source_id))


async def replace_file_source(
    tenant_id: str,
    user_id: str,
    source_id: str,
    file: UploadFile,
) -> dict:
    existing = get_source(tenant_id, source_id)
    if not existing:
        raise ValueError(f"Data source {source_id} not found")
    if existing.get("source_type") != "file":
        raise AppError(
            "data_source_not_replaceable",
            "Only file data sources can have their file replaced.",
            status_code=_REJECTED,
        )

    suffix = Path(file.filename or "").suffix.lower()
    _reject_unsupported_extension(suffix)

    content = await file.read()
    size_bytes = len(content)
    _reject_oversized_upload(size_bytes)

    from backend.storage import paths
    dst_dir = paths.dataset_dir(tenant_id, source_id)
    dst_dir.mkdir(parents=True, exist_ok=True)

    # Atomic write: save to .tmp first, verify, then swap — prevents data loss
    # if disk is full or write fails partway through.
    tmp_path = dst_dir / f"data{suffix}.tmp"
    try:
        with open(tmp_path, "wb") as f:
            f.write(content)
        if tmp_path.stat().st_size != size_bytes:
            raise ValueError("File write verification failed: size mismatch.")
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise

    for old in dst_dir.glob("data.*"):
        if old != tmp_path:
            old.unlink(missing_ok=True)

    file_path = dst_dir / f"data{suffix}"
    tmp_path.replace(file_path)

    # Eagerly count rows on replace so the UI shows a non-zero value immediately
    row_count = None
    col_count = None
    try:
        import io as _io
        if suffix == ".csv":
            row_count = max(0, content.count(b"\n") - 1)
            import csv as _csv
            header = content.split(b"\n", 1)[0].decode("utf-8", "replace")
            if header.strip():
                col_count = len(next(_csv.reader([header])))
        elif suffix in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content), read_only=True, data_only=True)
            row_count = sum((ws.max_row or 1) - 1 for ws in wb.worksheets)
            col_count = max((ws.max_column or 0) for ws in wb.worksheets) or None
            wb.close()
        elif suffix == ".parquet":
            import pyarrow.parquet as pq
            meta = pq.read_metadata(_io.BytesIO(content))
            row_count = meta.num_rows
            col_count = meta.num_columns
        elif suffix == ".json":
            import json as _json_mod
            data = _json_mod.loads(content)
            if isinstance(data, list):
                row_count = len(data)
                if data and isinstance(data[0], dict):
                    col_count = len(data[0])
    except Exception:
        pass

    execute(
        """UPDATE datasets
           SET original_filename=%s, file_type=%s, file_path=%s,
               size_bytes=%s, row_count=%s, column_count=%s,
               preview_cache=NULL, connection_status='connected', updated_at=NOW()
           WHERE id=%s AND tenant_id=%s""",
        (file.filename, suffix.lstrip("."), str(file_path), size_bytes, row_count, col_count, source_id, tenant_id),
    )
    return _public(get_source(tenant_id, source_id))


# ── SQL sources ────────────────────────────────────────────────────────────────

def create_sql_source(
    tenant_id: str,
    user_id: str,
    name: str,
    host: str,
    port: int,
    database: str,
    username: str,
    password: str,
    engine: str,
    description: Optional[str] = None,
) -> dict:
    if engine not in SQL_ENGINES:
        raise ValueError(f"Unsupported engine '{engine}'. Options: {sorted(SQL_ENGINES)}")

    source_id = generate_id("ds")
    sql_config = {
        "host": host,
        "port": port,
        "database": database,
        "username": username,
        "password_enc": _enc(password),
        "engine": engine,
    }
    execute(
        """INSERT INTO datasets
           (id, tenant_id, name, description, file_type, source_type,
            connection_status, sql_config, uploaded_by, uploaded_at, updated_at)
           VALUES (%s,%s,%s,%s,%s,'sql','pending',%s,%s,NOW(),NOW())""",
        (source_id, tenant_id, name, description, engine, _json(sql_config), user_id),
    )
    return _public(get_source(tenant_id, source_id))


def update_sql_config(
    tenant_id: str,
    source_id: str,
    host: str,
    port: int,
    database: str,
    username: str,
    password: Optional[str],
    engine: str,
) -> dict:
    existing = get_source(tenant_id, source_id)
    if not existing:
        raise ValueError(f"Data source {source_id} not found")

    old_cfg = existing.get("sql_config") or {}
    # Preserve password if not provided
    if password:
        enc_pw = _enc(password)
    else:
        enc_pw = old_cfg.get("password_enc", "")

    sql_config = {
        "host": host,
        "port": port,
        "database": database,
        "username": username,
        "password_enc": enc_pw,
        "engine": engine,
    }
    execute(
        """UPDATE datasets
           SET sql_config=%s, file_type=%s, connection_status='pending', updated_at=NOW()
           WHERE id=%s AND tenant_id=%s""",
        (_json(sql_config), engine, source_id, tenant_id),
    )
    return _public(get_source(tenant_id, source_id))


def test_sql_connection(tenant_id: str, source_id: str) -> dict:
    existing = get_source(tenant_id, source_id)
    if not existing:
        raise ValueError(f"Data source {source_id} not found")
    cfg = existing.get("sql_config") or {}
    try:
        import sqlalchemy
        engine = _make_sql_engine(cfg, statement_timeout_ms=5_000)
        with engine.connect() as conn:
            conn.execute(sqlalchemy.text("SELECT 1"))
        execute(
            "UPDATE datasets SET connection_status='connected', updated_at=NOW() WHERE id=%s AND tenant_id=%s",
            (source_id, tenant_id),
        )
        return {"ok": True, "status": "connected"}
    except Exception as e:
        execute(
            "UPDATE datasets SET connection_status='error', updated_at=NOW() WHERE id=%s AND tenant_id=%s",
            (source_id, tenant_id),
        )
        return {"ok": False, "status": "error", "error": str(e)}


def execute_sql_query(tenant_id: str, source_id: str, sql: str, limit: int = 500) -> dict:
    existing = get_source(tenant_id, source_id)
    if not existing:
        raise ValueError(f"Data source {source_id} not found")
    if existing.get("connection_status") != "connected":
        raise AppError(
            "data_source_not_connected",
            "This data source is not connected. Test the connection first.",
            status_code=_REJECTED,
        )
    cfg = existing.get("sql_config") or {}
    try:
        import sqlalchemy
        engine = _make_sql_engine(cfg, statement_timeout_ms=30_000)
        with engine.connect() as conn:
            result = conn.execute(sqlalchemy.text(sql))
            rows = result.fetchmany(limit)
            columns = list(result.keys())
        data = [dict(zip(columns, row)) for row in rows]
        return {
            "columns": columns,
            "rows": data,
            "row_count": len(data),
            "truncated": len(data) >= limit,
        }
    except AppError:
        raise
    except Exception as e:
        # The user's own SQL failed. The driver's text is the only useful part
        # and cannot be localized, so it travels as a param the frontend drops
        # into its own sentence.
        raise AppError(
            "sql_query_failed",
            f"Query failed: {e}",
            status_code=_REJECTED,
            params={"reason": str(e)},
        )


def materialize_sql_source(
    tenant_id: str,
    user_id: str,
    source_id: str,
    sql: Optional[str] = None,
    name: Optional[str] = None,
) -> dict:
    """Run a SQL source's query and snapshot the result as a NEW CSV dataset
    (parent_id = the SQL source). From that dataset on, the pipeline treats it
    exactly like an uploaded CSV — the wizard, training and freshness never
    need to know SQL was involved. A snapshot (not a live query) is deliberate:
    a forecast must be reproducible against the data it actually trained on.

    Streams the result in batches so the row cap bounds memory, and refuses —
    rather than silently truncates — when the result exceeds it."""
    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")
    if src.get("source_type") != "sql":
        raise AppError(
            "data_source_not_sql",
            "Only SQL data sources can be materialized into a dataset.",
            status_code=_REJECTED,
        )
    if src.get("connection_status") != "connected":
        raise AppError(
            "data_source_not_connected",
            "This data source is not connected. Test the connection first.",
            status_code=_REJECTED,
        )
    query_sql = (sql or src.get("saved_query") or "").strip()
    if not query_sql:
        raise AppError(
            "sql_source_no_saved_query",
            "This SQL data source has no query to materialize yet.",
            status_code=_REJECTED,
        )

    import csv as _csv

    import sqlalchemy

    from backend.storage import paths
    from backend.utils.csv_safe import csv_safe

    max_rows = settings.sql_materialize_max_rows
    new_id = generate_id("ds")
    dst_dir = paths.dataset_dir(tenant_id, new_id)
    dst_dir.mkdir(parents=True, exist_ok=True)
    file_path = dst_dir / "data.csv"
    tmp_path = dst_dir / "data.csv.tmp"

    row_count = 0
    columns: list[str] = []
    try:
        engine = _make_sql_engine(cfg=src.get("sql_config") or {}, statement_timeout_ms=120_000)
        with engine.connect() as conn:
            result = conn.execution_options(stream_results=True).execute(
                sqlalchemy.text(query_sql)
            )
            columns = [str(c) for c in result.keys()]
            with open(tmp_path, "w", newline="", encoding="utf-8") as f:
                writer = _csv.writer(f)
                writer.writerow([csv_safe(c) for c in columns])
                while True:
                    batch = result.fetchmany(10_000)
                    if not batch:
                        break
                    row_count += len(batch)
                    if row_count > max_rows:
                        raise AppError(
                            "sql_result_too_large",
                            f"The query returned more than {max_rows} rows, the "
                            "materialization limit. Narrow it with WHERE or LIMIT.",
                            status_code=_REJECTED,
                            params={"max_rows": max_rows},
                        )
                    for r in batch:
                        writer.writerow([_safe_cell(v) for v in r])
    except AppError:
        tmp_path.unlink(missing_ok=True)
        raise
    except Exception as e:
        tmp_path.unlink(missing_ok=True)
        raise AppError(
            "sql_query_failed",
            f"Query failed: {e}",
            status_code=_REJECTED,
            params={"reason": str(e)},
        )

    if row_count == 0:
        tmp_path.unlink(missing_ok=True)
        raise AppError(
            "sql_result_empty",
            "The query returned no rows — there is nothing to materialize.",
            status_code=_REJECTED,
        )

    size_bytes = tmp_path.stat().st_size
    try:
        # Same plan quota an uploaded file of this size would face.
        from backend.datasets.service import _enforce_dataset_size
        _enforce_dataset_size(tenant_id, size_bytes)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    tmp_path.replace(file_path)

    # English fallback on purpose: this becomes the dataset's persisted NAME
    # and backend logic must not hardcode Spanish — the UI sends a localized
    # name (same convention as save_edited_as_new).
    display_name = name.strip() if (name and name.strip()) else f"{src.get('name')} (SQL)"

    execute(
        """INSERT INTO datasets
           (id, tenant_id, name, description, original_filename, file_type, file_path,
            size_bytes, row_count, column_count, source_type, connection_status,
            parent_id, uploaded_by, uploaded_at, updated_at)
           VALUES (%s,%s,%s,%s,%s,'csv',%s,%s,%s,%s,'file','connected',%s,%s,NOW(),NOW())""",
        (
            new_id, tenant_id, display_name, src.get("description"),
            f"{display_name}.csv", str(file_path),
            size_bytes, row_count, len(columns), source_id, user_id,
        ),
    )
    # Remember the query that produced the snapshot, so re-materializing after
    # fresh rows land in the customer's DB is one click, not a rewrite.
    execute(
        "UPDATE datasets SET saved_query=%s, updated_at=NOW() WHERE id=%s AND tenant_id=%s",
        (query_sql, source_id, tenant_id),
    )
    return _public(get_source(tenant_id, new_id))


def export_sql_query_xlsx(tenant_id: str, source_id: str, sql: Optional[str] = None) -> bytes:
    """Run a SQL source's query and return the FULL result as an .xlsx workbook.

    Same streaming fetch and row ceiling as materialize (refuse, never
    truncate). Text cells go through the formula-injection guard — Excel
    executes a leading '=' even more eagerly than a re-imported CSV does."""
    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")
    if src.get("source_type") != "sql":
        raise AppError(
            "data_source_not_sql",
            "Only SQL data sources can be exported this way.",
            status_code=_REJECTED,
        )
    if src.get("connection_status") != "connected":
        raise AppError(
            "data_source_not_connected",
            "This data source is not connected. Test the connection first.",
            status_code=_REJECTED,
        )
    query_sql = (sql or src.get("saved_query") or "").strip()
    if not query_sql:
        raise AppError(
            "sql_source_no_saved_query",
            "This SQL data source has no query to export yet.",
            status_code=_REJECTED,
        )

    import io

    import sqlalchemy
    from openpyxl import Workbook

    from backend.utils.csv_safe import csv_safe

    max_rows = settings.sql_materialize_max_rows
    row_count = 0
    try:
        engine = _make_sql_engine(cfg=src.get("sql_config") or {}, statement_timeout_ms=120_000)
        with engine.connect() as conn:
            result = conn.execution_options(stream_results=True).execute(
                sqlalchemy.text(query_sql)
            )
            columns = [str(c) for c in result.keys()]
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="data")
            ws.append([csv_safe(c) for c in columns])
            while True:
                batch = result.fetchmany(10_000)
                if not batch:
                    break
                row_count += len(batch)
                if row_count > max_rows:
                    raise AppError(
                        "sql_result_too_large",
                        f"The query returned more than {max_rows} rows, the "
                        "export limit. Narrow it with WHERE or LIMIT.",
                        status_code=_REJECTED,
                        params={"max_rows": max_rows},
                    )
                for r in batch:
                    ws.append([_safe_cell(v) for v in r])
    except AppError:
        raise
    except Exception as e:
        raise AppError(
            "sql_query_failed",
            f"Query failed: {e}",
            status_code=_REJECTED,
            params={"reason": str(e)},
        )

    if row_count == 0:
        raise AppError(
            "sql_result_empty",
            "The query returned no rows — there is nothing to export.",
            status_code=_REJECTED,
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_sql_query(tenant_id: str, source_id: str, sql: str) -> dict:
    execute(
        "UPDATE datasets SET saved_query=%s, updated_at=NOW() WHERE id=%s AND tenant_id=%s",
        (sql, source_id, tenant_id),
    )
    return _public(get_source(tenant_id, source_id))


def _build_conn_str(cfg: dict) -> str:
    host = cfg.get("host", "localhost")
    port = cfg.get("port", 5432)
    database = cfg.get("database", "")
    username = cfg.get("username", "")
    password = _dec(cfg.get("password_enc", "")) if cfg.get("password_enc") else ""
    engine = cfg.get("engine", "postgresql")
    from urllib.parse import quote_plus
    pw_enc = quote_plus(password)
    user_enc = quote_plus(username)
    drivers = {
        "postgresql": "postgresql+psycopg2",
        "mysql": "mysql+pymysql",
        "mssql": "mssql+pyodbc",
        "oracle": "oracle+cx_oracle",
    }
    driver = drivers.get(engine, "postgresql+psycopg2")
    url = f"{driver}://{user_enc}:{pw_enc}@{host}:{port}/{database}"
    if engine == "mssql":
        # pyodbc needs an explicit ODBC driver name; pick the newest installed.
        # TrustServerCertificate matches Driver 18's new encrypt-by-default,
        # which otherwise rejects the self-signed certs typical of on-prem
        # SQL Servers at SMB distributors.
        url += f"?driver={quote_plus(_best_mssql_odbc_driver())}&TrustServerCertificate=yes"
    return url


def _best_mssql_odbc_driver() -> str:
    """Newest SQL Server ODBC driver installed on this host, falling back to
    the legacy 'SQL Server' driver that ships with Windows."""
    try:
        import pyodbc
        installed = [d for d in pyodbc.drivers() if "SQL Server" in d]
        for preferred in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"):
            if preferred in installed:
                return preferred
        if installed:
            return installed[-1]
    except Exception:
        pass
    return "ODBC Driver 18 for SQL Server"


# ── Preview ────────────────────────────────────────────────────────────────────

def get_preview(tenant_id: str, source_id: str, rows: int = 100, sheet: Optional[str] = None) -> dict:
    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")

    if src.get("source_type") == "sql":
        saved_q = src.get("saved_query")
        if not saved_q:
            return {"columns": [], "rows": [], "row_count": 0, "sheets": None}
        return execute_sql_query(tenant_id, source_id, saved_q, limit=rows)

    # File source
    file_path = src.get("file_path")
    if not file_path or not Path(file_path).exists():
        raise AppError(
            "data_source_file_missing",
            "The file is no longer on disk. Upload it again.",
            status_code=_REJECTED,
        )

    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx", ".xls"):
        _check_file_size(file_path)

    from backend.dataframes.io import dataset_preview
    preview = dataset_preview(file_path, rows, sheet=sheet)
    columns = preview["columns"]
    data_rows = preview["rows"]
    sheets = preview["sheets"]
    total_rows = preview["total_rows"]

    # Update stats if not set — the preview already read the full row count for
    # Excel/JSON/Parquet; CSV counts lines without loading the whole file.
    # Also backfill column_count when missing (older CSV uploads stored only a
    # row count, so the metadata card showed "COLUMNAS —" despite the preview
    # rendering all columns).
    if not src.get("row_count") or not src.get("column_count"):
        try:
            if total_rows is None:
                if src.get("row_count"):
                    total_rows = src["row_count"]
                else:
                    with open(file_path, "r", encoding="utf-8", errors="replace") as _f:
                        total_rows = sum(1 for _ in _f) - 1
            execute(
                "UPDATE datasets SET row_count=%s, column_count=%s, updated_at=NOW() WHERE id=%s AND tenant_id=%s",
                (total_rows, len(columns), source_id, tenant_id),
            )
        except Exception as _e:
            log.warning("[preview] failed to update stats for source=%s: %s", source_id, _e)

    return {
        "columns": columns,
        "rows": data_rows,
        "row_count": len(data_rows),
        "sheets": sheets,
        "active_sheet": sheet or (sheets[0] if sheets else None),
        "truncated": len(data_rows) >= rows,
    }


# ── In-app editor ────────────────────────────────────────────────────────────

def _safe_cell(value):
    """CSV formula-injection guard for editor cells that preserves numbers.

    A genuine number — including negatives like ``-5`` and signed/scientific
    forms — is never a spreadsheet formula, so it must NOT be quote-prefixed:
    doing so (``-5`` -> ``'-5``) would turn the value into text on re-read and
    corrupt the column. Numeric values pass through unchanged; only non-numeric
    text goes through the formula guard (``csv_safe``)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):   # bool is an int subclass; harmless here
        return value
    text = str(value)
    try:
        float(text)
        return text                        # numeric string ("-5", "+3", "1e3")
    except ValueError:
        from backend.utils.csv_safe import csv_safe
        return csv_safe(value)


def _not_editable() -> AppError:
    """Both editor entry points (load + save-as-new) refuse a SQL source with
    the same code, so the UI shows one sentence for one rule."""
    return AppError(
        "data_source_not_editable",
        "Only file data sources can be edited online.",
        status_code=_REJECTED,
    )


def load_editable_table(tenant_id: str, source_id: str) -> dict:
    """Load a file dataset's full table for in-browser editing. Enforces the
    size guard from the STORED row_count/size_bytes before touching the file, so
    an over-threshold dataset is never read into memory. SQL sources rejected."""
    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")
    if src.get("source_type") == "sql":
        raise _not_editable()

    max_rows = settings.dataset_editor_max_rows
    max_bytes = settings.dataset_editor_max_mb * 1024 * 1024
    row_count = src.get("row_count")
    size_bytes = src.get("size_bytes")
    if row_count is not None and row_count > max_rows:
        raise AppError(
            "dataset_too_many_rows_to_edit",
            f"This dataset has {row_count} rows, over the {max_rows}-row online "
            "editing limit. Edit it offline and upload it again.",
            status_code=_REJECTED,
            params={"rows": row_count, "max_rows": max_rows},
        )
    if size_bytes is not None and size_bytes > max_bytes:
        raise AppError(
            "dataset_too_large_to_edit",
            f"This dataset is {_mb(size_bytes)} MB, over the "
            f"{settings.dataset_editor_max_mb} MB online editing limit. "
            "Edit it offline and upload it again.",
            status_code=_REJECTED,
            params={"size_mb": _mb(size_bytes), "max_mb": settings.dataset_editor_max_mb},
        )

    file_path = src.get("file_path")
    if not file_path or not Path(file_path).exists():
        raise AppError(
            "data_source_file_missing",
            "The file is no longer on disk. Upload it again.",
            status_code=_REJECTED,
        )

    from backend.dataframes.io import read_table
    return read_table(file_path)


def save_edited_as_new(
    tenant_id: str,
    user_id: str,
    source_id: str,
    name: Optional[str],
    columns: list[str],
    rows: list[dict],
) -> dict:
    """Write the edited table as a NEW CSV dataset with `parent_id` = source id.
    The original dataset row and file are never touched. Every cell (and header)
    is passed through the CSV formula-injection guard before it is written."""
    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")
    if src.get("source_type") == "sql":
        raise _not_editable()
    if not columns or not isinstance(columns, list):
        raise ValueError("At least one column is required")

    from backend.utils.csv_safe import csv_safe
    from backend.dataframes.io import write_rows
    from backend.storage import paths

    safe_columns = [csv_safe(c) for c in columns]
    safe_rows = [
        {safe_columns[i]: _safe_cell(row.get(columns[i])) for i in range(len(columns))}
        for row in rows
    ]

    new_id = generate_id("ds")
    dst_dir = paths.dataset_dir(tenant_id, new_id)
    dst_dir.mkdir(parents=True, exist_ok=True)

    # Atomic write: write to .tmp, verify, then swap — a partial write never
    # leaves a corrupt dataset (mirrors replace_file_source).
    file_path = dst_dir / "data.csv"
    tmp_path = dst_dir / "data.csv.tmp"
    try:
        write_rows(str(tmp_path), safe_columns, safe_rows, fmt="csv")
        if not tmp_path.exists() or tmp_path.stat().st_size == 0:
            raise ValueError("File write verification failed.")
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise
    tmp_path.replace(file_path)

    size_bytes = file_path.stat().st_size
    row_count = len(rows)
    col_count = len(columns)
    # English fallback on purpose: this string becomes the dataset's persisted
    # NAME, and a Spanish literal in backend logic breaks CLAUDE.md. The UI
    # always sends `name` (defaulted from i18n, so a Spanish user still gets
    # "… (editado)"); this only fires for a direct API caller.
    display_name = name.strip() if (name and name.strip()) else f"{src.get('name')} (edited)"

    execute(
        """INSERT INTO datasets
           (id, tenant_id, name, description, original_filename, file_type, file_path,
            size_bytes, row_count, column_count, source_type, connection_status,
            parent_id, uploaded_by, uploaded_at, updated_at)
           VALUES (%s,%s,%s,%s,%s,'csv',%s,%s,%s,%s,'file','connected',%s,%s,NOW(),NOW())""",
        (
            new_id, tenant_id, display_name, src.get("description"),
            f"{display_name}.csv", str(file_path),
            size_bytes, row_count, col_count, source_id, user_id,
        ),
    )
    return _public(get_source(tenant_id, new_id))


# ── CRUD ───────────────────────────────────────────────────────────────────────

def get_source(tenant_id: str, source_id: str) -> Optional[dict]:
    return query_one(
        "SELECT * FROM datasets WHERE id=%s AND tenant_id=%s",
        (source_id, tenant_id),
    )


def list_sources(tenant_id: str, skip: int = 0, limit: int = 50) -> list[dict]:
    rows = query(
        "SELECT * FROM datasets WHERE tenant_id=%s ORDER BY updated_at DESC NULLS LAST, uploaded_at DESC LIMIT %s OFFSET %s",
        (tenant_id, limit, skip),
    )
    return [_public(r) for r in rows]


def count_sources(tenant_id: str) -> int:
    row = query_one("SELECT COUNT(*) AS cnt FROM datasets WHERE tenant_id=%s", (tenant_id,))
    return row["cnt"] if row else 0


def delete_source(tenant_id: str, source_id: str) -> None:
    src = get_source(tenant_id, source_id)
    if not src:
        return
    # Remove file if exists
    if src.get("file_path"):
        p = Path(src["file_path"])
        if p.exists():
            p.unlink(missing_ok=True)
        if p.parent.exists() and not any(p.parent.iterdir()):
            p.parent.rmdir()
    execute("DELETE FROM datasets WHERE id=%s AND tenant_id=%s", (source_id, tenant_id))


def rename_source(tenant_id: str, source_id: str, name: str, description: Optional[str] = None) -> dict:
    execute(
        "UPDATE datasets SET name=%s, description=%s, updated_at=NOW() WHERE id=%s AND tenant_id=%s",
        (name, description, source_id, tenant_id),
    )
    return _public(get_source(tenant_id, source_id))


# ── Analysis helpers ───────────────────────────────────────────────────────────

def load_dataframe(tenant_id: str, source_id: str, sheet: Optional[str] = None, max_rows: int = 50_000):
    """Load DataFrame for analysis, capped at max_rows to bound memory usage."""
    from backend.dataframes.io import read_dataframe, dataframe_from_records

    src = get_source(tenant_id, source_id)
    if not src:
        raise ValueError(f"Data source {source_id} not found")

    if src.get("source_type") == "sql":
        saved_q = src.get("saved_query")
        if not saved_q:
            raise AppError(
                "sql_source_no_saved_query",
                "This SQL data source has no saved query to analyze yet.",
                status_code=_REJECTED,
            )
        cfg = src.get("sql_config") or {}
        import sqlalchemy
        eng = _make_sql_engine(cfg, statement_timeout_ms=60_000)
        with eng.connect() as conn:
            result = conn.execute(sqlalchemy.text(saved_q))
            rows = result.fetchmany(max_rows)
            cols = list(result.keys())
        return dataframe_from_records(rows, cols)

    file_path = src.get("file_path")
    if not file_path or not Path(file_path).exists():
        raise AppError(
            "data_source_file_missing",
            "The file is no longer on disk. Upload it again.",
            status_code=_REJECTED,
        )

    suffix = Path(file_path).suffix.lower()
    if suffix in (".xlsx", ".xls"):
        _check_file_size(file_path)
        return read_dataframe(file_path, sheet=sheet, nrows=max_rows)
    elif suffix == ".json":
        return read_dataframe(file_path, nrows=max_rows)
    else:
        return read_dataframe(file_path, nrows=max_rows)


def detect_columns(df) -> dict:
    """Heuristic auto-detection of date, target, and group columns."""
    cols = list(df.columns)
    lower = {c.lower(): c for c in cols}

    date_hints    = ["date", "fecha", "dt", "time", "period", "week", "month", "year", "timestamp", "dia", "semana", "mes"]
    target_hints  = ["sales", "ventas", "demand", "qty", "quantity", "units", "target", "value", "amount", "revenue", "uds", "importe", "venta"]
    sku_hints     = ["sku", "product", "item", "grupo", "group", "category", "categoria", "codigo", "code", "product_id", "item_id"]

    def find(hints):
        for h in hints:
            for lc, orig in lower.items():
                if h in lc:
                    return orig
        return None

    date_col = find(date_hints)
    if not date_col:
        for c in cols:
            if str(df[c].dtype).startswith("datetime"):
                date_col = c
                break

    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    target_col = find(target_hints)
    if not target_col:
        for c in numeric_cols:
            if c != date_col and "id" not in c.lower():
                target_col = c
                break
    if not target_col and numeric_cols:
        target_col = numeric_cols[0]

    str_cols = df.select_dtypes(include=["object", "category"]).columns.tolist()
    sku_col = find(sku_hints)
    if not sku_col:
        for c in str_cols:
            if c != date_col:
                sku_col = c
                break

    return {"date_col": date_col, "target_col": target_col, "sku_col": sku_col}
